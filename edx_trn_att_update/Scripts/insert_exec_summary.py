#!/usr/bin/env python3
# Adds (or replaces) an Executive_Summary sheet to a workbook
# v1 103125_1653
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path
import argparse
from datetime import datetime

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-xlsx", required=True)
    args = ap.parse_args()

    p = Path(args.in_xlsx)
    wb = load_workbook(p)
    name = "Executive_Summary"
    if name in wb.sheetnames:
        ws = wb[name]
        for r in ws["A1:A400"]:
            for c in r: c.value = None
    else:
        ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Executive Summary — UAP Training Attainment"
    ws["A1"].font = Font(bold=True, size=14)

    sections = [
        ("Purpose","This workbook gives leaders a consistent view of training engagement and module performance. It blends:\n• UAP Attempts (SR04-Trn_Att)\n• HR Master (Terms excluded via DNR)\n• Module Classification (Entity → SOP → SubSOP)\nUse slicers to focus by team, location, or function."),
        ("Tabs","• Report_Card — per-user × module\n• Last_Active — recent activity with aging buckets\n• Module_Risk — conversion %, attempts-to-pass, score stats\n• Run_Metadata — inputs & versions\n• Source_Columns — audit of input headers"),
        ("How to use slicers","1) Click a slicer item to filter. 2) Ctrl-click for multi. 3) Use Timeline for dates. 4) Clear via the funnel-with-X icon."),
        ("Contacts / ownership","• Report Owner: Subi Rajagopalan\n• Data Stewards: UAP Admin, HR\n• Build Date: (auto-stamped in Run_Metadata)"),
    ]
    r = 3
    for h, body in sections:
        ws[f"A{r}"] = h; ws[f"A{r}"].font = Font(bold=True, size=11); r += 1
        for line in body.split("\n"):
            ws[f"A{r}"] = line
            ws[f"A{r}"].alignment = Alignment(wrap_text=True)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 120
    wb.save(p)
    print(str(p))

if __name__ == "__main__":
    main()
