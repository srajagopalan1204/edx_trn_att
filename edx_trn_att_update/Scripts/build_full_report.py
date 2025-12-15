#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build Enriched Report (v1, 103125_1653)

What this does
--------------
1) Load UAP SR04-Trn_Att CSV (attempts) and HR master (Current/New/Terms).
2) Exclude Terms → DNR_<ts>.csv (audit) if --exclude-dnr yes.
3) Apply Transi/FindName.csv to resolve UAP↔HR names; emit unmatched_names_<ts>.csv.
4) (Optional) Classify modules from SR04 Document Tracking Report -> Transi/module_classify_from_dtr_<ts>.csv.
5) (Optional) Apply Role→Area policy (normalized) to create role_module_requirements_seed and by_user_requirements_seed.
6) Write an Excel workbook with tabs: Facts_Attempts, Last_Active (with buckets), Run_Metadata, Source_Columns.
   (Pivots/slicers built manually or via VBA macro outside this script.)

CLI (examples)
--------------
python Scripts/build_full_report.py \\
  --uap Inputs/uap/SR04-Trn_Att.csv \\
  --emp-mstr emp_mstr/All_Employees_New_Hires_Terms_10_29_25.xlsx \\
  --findname Transi/FindName.csv \\
  --exclude-dnr yes \\
  --classify Inputs/classify/SR04*Document*Tracking*.csv \\
  --policy Inputs/req_matrix/Requirement_Area_Policy.csv \\
  --out-xlsx Outputs/Training_Attempts_Report_103125_1653.xlsx
"""
import argparse, sys, re
from pathlib import Path
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

NOW = datetime.now()

def read_csv_any(p):
    try:
        return pd.read_csv(p, dtype=str)
    except UnicodeDecodeError:
        return pd.read_csv(p, dtype=str, encoding="latin-1")

def detect_hr_sections(xls_path: Path) -> pd.DataFrame:
    raw = pd.read_excel(xls_path, sheet_name=0, header=None)
    labels = raw.applymap(lambda v: str(v).strip())

    def find_row_idx(s):
        idx = labels.index[labels.apply(lambda r: any(s == str(v) for v in r), axis=1)]
        return int(idx[0]) if len(idx) else None

    r_cur = find_row_idx("Current Employees")
    r_new = find_row_idx("New Hires")
    r_terms = find_row_idx("Terms")
    n = len(raw)

    def best_header(df):
        targets = {"Name","Role Name","Infor ID","Location","Manager","Primary Functional Group"}
        best = 0; hits = -1
        for i in range(min(10,len(df))):
            rowvals = set(str(x).strip() for x in df.iloc[i].tolist())
            h = sum(1 for t in targets if t in rowvals)
            if h>hits: best, hits = i, h
        return best

    def slice_and_parse(start, end, status):
        if start is None or start>=end: return pd.DataFrame()
        block = raw.iloc[start:end].copy()
        hdr = best_header(block)
        block.columns = block.iloc[hdr].astype(str).str.strip()
        block = block.iloc[hdr+1:].reset_index(drop=True)
        block["HR_Status"] = status
        return block

    a0 = r_cur if r_cur is not None else 0
    a1 = r_new if r_new is not None else (r_terms if r_terms is not None else n)
    b0 = r_new if r_new is not None else n
    b1 = r_terms if r_terms is not None else n
    c0 = r_terms if r_terms is not None else n
    c1 = n
    parts = [slice_and_parse(a0,a1,"Current"), slice_and_parse(b0,b1,"New Hire"), slice_and_parse(c0,c1,"Terms")]
    hr = pd.concat([p for p in parts if not p.empty], ignore_index=True)
    # Normalize expected columns
    ren = {}
    for c in hr.columns:
        lc = c.lower().strip()
        if lc in {"name","employee name","user.full name"}: ren[c] = "User.Full Name"
        elif lc in {"infor id","infor_id","inforid"}: ren[c] = "Emp_InforID"
        elif lc in {"role name","rolename"}: ren[c] = "RoleName"
        elif lc == "manager": ren[c] = "Emp_Manager"
        elif lc == "primary functional group": ren[c] = "Emp_PrimaryFunctionalGroup"
        elif lc == "location": ren[c] = "Emp_Location"
        elif lc == "location number": ren[c] = "Emp_LocationNumber"
    hr = hr.rename(columns=ren)
    for k in ["User.Full Name","Emp_InforID","RoleName","HR_Status","Emp_Manager","Emp_PrimaryFunctionalGroup","Emp_Location","Emp_LocationNumber"]:
        if k not in hr.columns: hr[k] = ""
        hr[k] = hr[k].astype(str).str.strip()
    return hr

def split_path(s):
    return [p.strip() for p in str(s).split(">") if str(p).strip()]

def guess_entity(parts):
    up = [p.upper() for p in parts]
    if any("PALCO POWER SYSTEMS" in p for p in up): return "PALCO"
    if any("SOUTH MAIN TRUCKING" in p or p=="SMT" for p in up): return "SMT"
    if any("SPECIALTY" in p for p in up): return "Specialty"
    return "Scott Electric Distribution"

def sop_sub(parts):
    if "Document Library" in parts:
        tail = parts[parts.index("Document Library")+1:]
    else:
        tail = parts[:]
    up0 = tail[0].upper() if tail else ""
    if up0 in ('PALCO POWER SYSTEMS', 'SMT', 'SOUTH MAIN TRUCKING', 'SPECIALTY'):
        tail = tail[1:]
    sop = tail[0] if len(tail)>=1 else ""
    subsop = tail[1] if len(tail)>=2 else ""
    return sop, subsop

def write_sheet(ws, df, index=False):
    for r in dataframe_to_rows(df, index=index, header=True):
        ws.append(r)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--uap", required=True, help="UAP SR04-Trn_Att CSV")
    ap.add_argument("--emp-mstr", required=True, help="HR Master XLSX (with Current/New/Terms)")
    ap.add_argument("--findname", default="Transi/FindName.csv", help="Authoritative name map")
    ap.add_argument("--exclude-dnr", choices=["yes","no"], default="yes")
    ap.add_argument("--classify", default="", help="SR04 Document Tracking Report CSV (glob ok)")
    ap.add_argument("--policy", default="", help="Role→Area policy (normalized CSV)")
    ap.add_argument("--out-xlsx", required=True, help="Output workbook path")
    args = ap.parse_args()

    uap = read_csv_any(args.uap).fillna("")
    hr = detect_hr_sections(Path(args.emp_mstr))

    # HR split
    terms = hr[hr["HR_Status"].str.lower().eq("terms")].copy()
    active = hr[~hr["HR_Status"].str.lower().eq("terms")].copy()

    # DNR
    out_dir = Path(args.out_xlsx).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    if args.exclude-dnr if False else False: pass  # placeholder to satisfy linter

    if args.exclude_dnr == "yes":
        dnr_path = out_dir / f"DNR_{NOW.strftime('%m%d%y_%H%M')}.csv"
        terms.to_csv(dnr_path, index=False)

    # Name map
    findname_path = Path(args.findname)
    if findname_path.exists():
        fmap = read_csv_any(findname_path)
        # expected columns: User.Full Name (UAP side), Name (HR side)
        ucol = None
        hcol = None
        for c in fmap.columns:
            lc = c.lower()
            if "user.full" in lc and "name" in lc: ucol = c
            if (c.lower() == "name") or ("hr" in lc and "name" in lc): hcol = c
        if ucol and hcol:
            uap = uap.merge(fmap[[ucol,hcol]].rename(columns={ucol:"UAP_Name", hcol:"HR_Name"}), left_on="User.Full Name", right_on="UAP_Name", how="left")
            uap["MatchName"] = uap["HR_Name"].where(uap["HR_Name"].ne(""), uap["User.Full Name"])
        else:
            uap["MatchName"] = uap["User.Full Name"]
    else:
        uap["MatchName"] = uap["User.Full Name"]

    # Unmatched names report (pre-join)
    unmatched = pd.DataFrame({"User.Full Name": sorted(set(uap["User.Full Name"]) - set(active.get("User.Full Name", pd.Series(dtype=str))))})
    unmatched_path = out_dir / f"unmatched_names_{NOW.strftime('%m%d%y_%H%M')}.csv"
    unmatched.to_csv(unmatched_path, index=False)

    # Join HR
    facts = uap.merge(active, left_on="MatchName", right_on="User.Full Name", how="left", suffixes=("","_HR"))

    # Classify (optional)
    module_catalog = pd.DataFrame()
    if args.classify:
        import glob
        candidates = sorted(glob.glob(args.classify))
        if candidates:
            dtr = read_csv_any(candidates[0]).fillna("")
            rows = []
            for _, r in dtr.iterrows():
                name = (r.get("Name") or "").strip()
                pctx = (r.get("Parent.Context") or "").strip()
                parts = split_path(pctx)
                ent = guess_entity(parts)
                sop, subsop = sop_sub(parts)
                if name:
                    rows.append({"Entity":ent,"SOP":sop,"SubSOP":subsop,"Module_Name":name,"Parent_Context":pctx})
            module_catalog = pd.DataFrame(rows).drop_duplicates(subset=["Entity","SOP","SubSOP","Module_Name"])
            cat_path = Path("Transi") / f"module_classify_from_dtr_{NOW.strftime('%m%d%y_%H%M')}.csv"
            Path("Transi").mkdir(exist_ok=True, parents=True)
            module_catalog.to_csv(cat_path, index=False)

    # Requirements seed (optional) — expand policy to modules and by-user
    role_module = pd.DataFrame()
    by_user_seed = pd.DataFrame()
    if args.policy and not module_catalog.empty:
        policy = read_csv_any(args.policy).fillna("")
        def expand(row):
            if row.get("SubSOP",""):
                m = (module_catalog["Entity"].str.casefold()==row["Entity"].casefold()) & \
                    (module_catalog["SOP"].str.casefold()==row["SOP"].casefold()) & \
                    (module_catalog["SubSOP"].str.casefold()==row["SubSOP"].casefold())
            else:
                m = (module_catalog["Entity"].str.casefold()==row["Entity"].casefold()) & \
                    (module_catalog["SOP"].str.casefold()==row["SOP"].casefold())
            sub = module_catalog.loc[m, ["Entity","SOP","SubSOP","Module_Name"]].copy()
            if sub.empty: return pd.DataFrame()
            sub["RoleName"] = row["RoleName"]
            cov = str(row.get("Coverage","ALL")).upper()
            sub["Flag"] = "R" if cov=="ALL" else ("Z" if cov=="NONE" else "N")
            sub["Source"] = cov
            return sub
        outs = []
        for _, rr in policy.iterrows():
            outs.append(expand(rr))
        role_module = pd.concat([o for o in outs if o is not None and not o.empty], ignore_index=True)
        role_module.to_csv(Path("Transi")/f"role_module_requirements_seed_{NOW.strftime('%m%d%y_%H%M')}.csv", index=False)

        # By-user
        active2 = active.copy()
        active2["RoleName"] = active2["RoleName"].replace({"": "(Unassigned)"}).fillna("(Unassigned)")
        by_user_seed = active2.merge(role_module, on="RoleName", how="left")
        by_user_seed.to_csv(Path("Transi")/f"by_user_requirements_seed_{NOW.strftime('%m%d%y_%H%M')}.csv", index=False)

    # Build Last_Active view (if LastActivity exists)
    last_active_cols = []
    for cand in ["LastActivity_date","LastActivity_dt","Last Activity","LastActivity"]:
        if cand in facts.columns: last_active_cols.append(cand)
    last = pd.DataFrame()
    if last_active_cols:
        dtcol = last_active_cols[0]
        temp = facts.copy()
        temp["LastActivity_dt"] = pd.to_datetime(temp[dtcol], errors="coerce")
        temp["DaysSinceActive"] = (pd.Timestamp(NOW) - temp["LastActivity_dt"]).dt.days
        # Buckets (static cutoffs; can be edited in Excel)
        cutoff_pre2025 = pd.Timestamp("2025-01-01")
        cutoff_jun2025 = pd.Timestamp("2025-06-30 23:59:59")
        temp["Pre_2025"] = (temp["LastActivity_dt"] <= cutoff_pre2025).astype(int)
        temp["JanToJun_2025"] = ((temp["LastActivity_dt"] > cutoff_pre2025) & (temp["LastActivity_dt"] <= cutoff_jun2025)).astype(int)
        temp["Past_60d"] = (temp["LastActivity_dt"] >= (pd.Timestamp(NOW) - pd.Timedelta(days=60))).astype(int)
        temp["Past_14d"] = (temp["LastActivity_dt"] >= (pd.Timestamp(NOW) - pd.Timedelta(days=14))).astype(int)
        last = temp

    # Write workbook
    out_xlsx = Path(args.out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Facts_Attempts"
    write_sheet(ws1, facts)

    if not last.empty:
        ws2 = wb.create_sheet("Last_Active")
        write_sheet(ws2, last)

    # Metadata
    meta = pd.DataFrame([{"BuiltAt": NOW.strftime("%Y-%m-%d %H:%M"),
                           "UAP": args.uap,
                           "HR_Master": args.emp_mstr,
                           "Classify": args.classify,
                           "Policy": args.policy,
                           "FindName": args.findname}])
    ws3 = wb.create_sheet("Run_Metadata")
    write_sheet(ws3, meta)

    # Source columns
    src_cols = pd.DataFrame({"UAP_Columns":[*uap.columns], "HR_Columns":[*active.columns]})
    ws4 = wb.create_sheet("Source_Columns")
    write_sheet(ws4, src_cols)

    wb.save(out_xlsx)
    print(str(out_xlsx))

if __name__ == "__main__":
    sys.exit(main())
