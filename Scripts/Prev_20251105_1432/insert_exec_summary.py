#!/usr/bin/env python3
"""
insert_exec_summary.py

Create/replace an 'Exec_Summary' sheet in a report workbook produced by build_full_report.py.
Summarizes:
- Source files (from Run_Metadata)
- Row counts per tab (Report_Card, Exceptions, Last_Active, Module_Risk, Module_Risk_By_Policy, Requirement_Matrix)
- Distinct users/modules
- DNR size (optional --dnr)
- Latest unmatched list size (optional --unmatched-glob)

Requires: pandas, openpyxl
"""

import argparse
import glob
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def safe_read_sheet(xlsx: Path, sheet: str) -> pd.DataFrame:
    try:
        return pd.read_excel(xlsx, sheet_name=sheet)
    except Exception:
        return pd.DataFrame()

def autosize(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row, start=1):
            cell_str = "" if cell is None else str(cell)
            widths[i] = max(widths.get(i, 0), len(cell_str))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 80)

def main():
    ap = argparse.ArgumentParser(description="Insert or refresh an Exec_Summary sheet in the report workbook.")
    ap.add_argument("--in-xlsx", required=True, help="Path to the Excel workbook to modify.")
    ap.add_argument("--dnr", help="Path to DNR CSV (optional, to show current size).")
    ap.add_argument("--unmatched-glob", help="Glob for unmatched CSVs (e.g., Outputs/unmatched_names_*.csv). Optional.")
    args = ap.parse_args()

    xlsx = Path(args.in_xlsx)
    if not xlsx.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx}")

    # Read needed tabs
    rc   = safe_read_sheet(xlsx, "Report_Card")
    exc  = safe_read_sheet(xlsx, "Exceptions")
    la   = safe_read_sheet(xlsx, "Last_Active")
    mr   = safe_read_sheet(xlsx, "Module_Risk")
    mrp  = safe_read_sheet(xlsx, "Module_Risk_By_Policy")
    reqm = safe_read_sheet(xlsx, "Requirement_Matrix")
    meta = safe_read_sheet(xlsx, "Run_Metadata")

    # Compute summary stats
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    users_distinct = rc["User.Full Name"].nunique() if "User.Full Name" in rc.columns else 0
    mods_distinct  = rc["Version Name"].nunique()  if "Version Name"  in rc.columns else 0

    # DNR size
    dnr_size = ""
    if args.dnr:
        try:
            import csv
            import pandas as pd
            dnr_df = pd.read_csv(args.dnr)
            # Accept either column name variants for safety
            name_col = "User.Full Name" if "User.Full Name" in dnr_df.columns else (dnr_df.columns[0] if len(dnr_df.columns)>0 else None)
            dnr_size = dnr_df[name_col].dropna().shape[0] if name_col else ""
        except Exception as e:
            dnr_size = f"ERR: {e}"

    # Unmatched latest file + count
    unmatched_file = ""
    unmatched_count = ""
    if args.unmatched_glob:
        files = sorted(glob.glob(args.unmatched_glob))
        if files:
            unmatched_file = files[-1]
            try:
                um = pd.read_csv(unmatched_file)
                # If we dropped a helper column in main pipeline, ignore it
                if "User.Full Name" in um.columns:
                    unmatched_count = int(um["User.Full Name"].nunique())
                else:
                    unmatched_count = int(um.shape[0])
            except Exception as e:
                unmatched_count = f"ERR: {e}"

    # Extract metadata fields if present
    src_attempt = ""
    src_emp = ""
    gen_ts = ""
    if not meta.empty:
        src_attempt = meta.iloc[0].get("Source_AttemptCSV", "")
        src_emp     = meta.iloc[0].get("Source_EmployeeMaster", "")
        gen_ts      = meta.iloc[0].get("Generated_Timestamp", "")

    # Prepare summary rows
    rows = [
        ["Executive Summary", ""],
        ["Generated (local time)", now_str],
        ["Source - Attempts CSV", src_attempt],
        ["Source - Employee Master", src_emp],
        ["Generated timestamp (from run)", gen_ts],
        ["" , ""],
        ["Row counts per tab", ""],
        ["Report_Card", rc.shape[0]],
        ["Exceptions",  exc.shape[0]],
        ["Last_Active", la.shape[0]],
        ["Module_Risk", mr.shape[0]],
        ["Module_Risk_By_Policy", mrp.shape[0]],
        ["Requirement_Matrix (cells)", f"{reqm.shape[0]} x {reqm.shape[1]}"],
        ["" , ""],
        ["Distinct Users (Report_Card)", users_distinct],
        ["Distinct Modules (Report_Card)", mods_distinct],
        ["" , ""],
        ["DNR size (rows)", dnr_size],
        ["Latest unmatched file", unmatched_file],
        ["Latest unmatched unique names", unmatched_count],
    ]

    # Open workbook, drop old Exec_Summary if present, add new
    wb = load_workbook(filename=xlsx)
    if "Exec_Summary" in wb.sheetnames:
        ws = wb["Exec_Summary"]
        wb.remove(ws)
    ws = wb.create_sheet("Exec_Summary", 0)

    # Write rows
    for r_idx, (c1, c2) in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=c1)
        ws.cell(row=r_idx, column=2, value=c2)

    autosize(ws)
    wb.save(xlsx)
    print(f"[INFO] Exec_Summary refreshed in: {xlsx}")

if __name__ == "__main__":
    main()
